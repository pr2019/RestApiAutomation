import json
import requests
import openpyxl
import jsonpath


class Common:
    def __init__(self, FileName, SheetName):
        global vk
        global sh
        vk = openpyxl.load_workbook(FileName)
        # print(vk.sheetnames)
        sh = vk[SheetName]

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        col = sh.max_column
        return col

    def fetch_key_names(self):
        c = self.fetch_column_count()
        l1 = []
        for i in range(1, c+1):
            cell = sh.cell(row=1, column=i)
            l1.insert(i, cell.value)

        return l1

    def update_req_with_data(self, rowNum, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=rowNum, column=i)
            jsonRequest[keyList[i-1]] = cell.value

        return jsonRequest



