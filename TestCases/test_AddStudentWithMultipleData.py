import requests
import json
import jsonpath
import openpyxl
from .Library import Common


def test_add_multiple_students():
    obj = Common("https://thetestingworldapi.com/api/studentsDetails",
                 "C:\\Users\\prade\\PycharmProjects\\ApiTesting\\TestCases\\RequestJosn.json")
    api_url, json_request = obj.common_steps()
    vk = openpyxl.load_workbook("C:/Users/prade/PycharmProjects/ApiTesting/TestCases/AddStudentsAData.xlsx")
    sh = vk["Sheet1"]
    rows = sh.max_row

    for i in range(1, rows + 1):
        cell_f_name = sh.cell(row=i, column=1).value
        cell_m_name = sh.cell(row=i, column=2).value
        cell_l_name = sh.cell(row=i, column=3).value
        cell_dob = sh.cell(row=i, column=4).value

        json_request["first_name"] = cell_f_name
        json_request["middle_name"] = cell_m_name
        json_request["last_name"] = cell_l_name
        json_request["date_of_birth"] = cell_dob

        response = requests.post(api_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
